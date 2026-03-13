"""
summary_prompts.py — LLM 摘要 Prompt 模板
A: 小型商務 (2-4人) | B: 中型多方 (5-8人) | C: 大型跨部門 (10+人) | D: EMBA 課堂
"""

from pathlib import Path
from typing import List, Dict

# ============================================================
# Glossary: 鋁輪圈/製程 中日英術語對照
# - 來源：memory/鋁輪圈專有名詞.xlsx（由使用者提供）
# - 用途：只把『逐字稿中有出現的術語』挑出來塞進 prompt，避免 prompt 爆長
# ============================================================

_CANDIDATE_GLOSSARY_PATHS = [
    Path.home() / '.openclaw' / 'workspace' / 'memory' / '鋁輪圈專有名詞.xlsx',
    Path.home() / '.openclaw' / 'workspace' / 'memory' / 'wheel_glossary.xlsx',
]


def _load_wheel_glossary_rows(xlsx_path: Path) -> List[Dict[str, str]]:
    '''Load glossary rows from an xlsx file.

    Columns are flexible; we try to match headers containing:
      - 中文 / 英文 / 日文 / 說明(備註)
    '''
    try:
        from openpyxl import load_workbook
    except Exception:
        return []

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        return []

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(c).strip() if c is not None else '' for c in rows[0]]

    def find_col(keys):
        for k in keys:
            for i, h in enumerate(header):
                if k in h:
                    return i
        return None

    zh_i = find_col(['中文', '中', '繁'])
    en_i = find_col(['英文', '英', 'EN'])
    jp_i = find_col(['日文', '日', 'JP', '日本語'])
    note_i = find_col(['說明', '備註', '註', 'note'])

    out = []
    for r in rows[1:]:
        def get(i):
            if i is None or i >= len(r):
                return ''
            v = r[i]
            return str(v).strip() if v is not None else ''

        item = {
            'zh': get(zh_i),
            'en': get(en_i),
            'jp': get(jp_i),
            'note': get(note_i),
        }
        if any(item.values()):
            out.append(item)
    return out


def _pick_glossary_for_transcript(transcript_text: str, max_items: int = 60) -> List[Dict[str, str]]:
    xlsx = next((pp for pp in _CANDIDATE_GLOSSARY_PATHS if pp.exists()), None)
    if not xlsx:
        return []

    rows = _load_wheel_glossary_rows(xlsx)
    if not rows:
        return []

    t = transcript_text or ''
    picked = []
    for it in rows:
        keys = [it.get('jp', ''), it.get('en', ''), it.get('zh', '')]
        keys = [k for k in keys if k and len(k) >= 2]
        if keys and any(k in t for k in keys):
            picked.append(it)
            if len(picked) >= max_items:
                break
    return picked


def _render_glossary_md(items: List[Dict[str, str]]) -> str:
    if not items:
        return ''
    lines = [
        '## 鋁輪圈/製程術語對照（優先採用）',
        '- 若逐字稿出現下列術語，請優先使用這份對照做翻譯/統一用詞',
        '- 若沒有對照，保留原文並在括號補繁中即可（不要自造名詞）',
        '',
        '| 中文 | English | 日本語 | 說明 |',
        '|---|---|---|---|',
    ]
    for it in items:
        lines.append(f"| {it.get('zh','')} | {it.get('en','')} | {it.get('jp','')} | {it.get('note','')} |")
    return "\n".join(lines) + "\n\n"



def select_template(meeting_type, speaker_count, duration_min, override=None):
    if override and override.upper() in ("A", "B", "C", "D", "E", "R"):
        return override.upper()
    if meeting_type == "emba":
        return "D"
    if speaker_count > 8 or duration_min > 120:
        return "C"
    elif speaker_count >= 5 or duration_min >= 60:
        return "B"
    return "A"


def select_model(duration_min, user_pref=None):
    m = {"claude": "anthropic/claude-sonnet-4-6",
         "gpt": "openai/gpt-5.2",
         "gemini": "google/gemini-3.1-pro-preview"}
    if user_pref and user_pref.lower() in m:
        return m[user_pref.lower()]
    if duration_min > 120:
        # Final reduce step benefits from stronger reasoning; chunks are already cheap.
        return m["claude"]
    elif duration_min > 30:
        return m["claude"]
    return m["gemini"]


from typing import Optional

def select_chunk_model(duration_min: float, user_pref: Optional[str] = None) -> str:
    """Pick a high C/P model for per-chunk processing.

    Goal: speed + low cost + low rate-limit risk.

    Note: chunk step should NOT use expensive reasoning models.
    """
    # Prefer Flash-tier models for chunk summarization.
    if user_pref and user_pref.lower() in ("gpt", "claude"):
        # user explicitly wants expensive model; still keep chunks cheap.
        pass
    return "google/gemini-3-flash-preview"


def build_chunk_prompt(
    transcript_chunk: str,
    chunk_label: str,
    metadata: dict,
    speakers: dict,
    bilingual: bool = False,
):
    """Return (system_prompt, user_message) for chunk summarization.

    Output must be compact and merge-friendly.
    """

    sys = """你是會議/課堂逐字稿的『分段整理器』。

目標：把一小段逐字稿壓縮成『可合併』的結構化筆記，避免冗詞。
規則：
- 只輸出繁體中文（台灣用語），除非特定術語需要括號保留原文。
- 句子要短；能條列就不寫長段落。
- 不要重述逐字稿；只抓資訊密度高的內容。
- 只輸出下列章節，缺就省略。

輸出格式（Markdown）：
### 分段資訊
- 段落：{chunk_label}

### 重點
- ...

### 決策/結論
- ...

### 行動項目
- [ ] ...

### 關鍵 Q&A
- Q: ...
  - A: ...

### 名詞/術語
- 術語：中文（原文） — 1 句解釋
""".replace("{chunk_label}", chunk_label)

    usr = f"""以下是逐字稿分段（{chunk_label}）。

==== 逐字稿 ====
{transcript_chunk}
"""

    return sys, usr


def build_reduce_prompt(
    tmpl_system: str,
    metadata_block: str,
    chunk_notes_md: str,
    glossary_md: str = "",
):
    """Combine chunk notes into the final summary prompt."""
    sys = tmpl_system.format(metadata_block=metadata_block)
    usr = """你會收到『多段分段筆記』，請把它們整合成最終摘要。

要求：
- 保持模板的章節與格式。
- 去重、合併同義項。
- 行動項目要具體、可執行，必要時補充負責方/截止日（不知道就留空）。
- 若分段筆記互相矛盾，列在『待確認事項』。

==== 分段筆記（已壓縮） ====
{chunk_notes}

{glossary}
""".format(chunk_notes=chunk_notes_md, glossary=glossary_md or "")
    return sys, usr


_LANG = """
## 語言處理（中文單語輸出）
- 全文只用**繁體中文（台灣用語）**撰寫（章節標題、內容、Q&A 都是中文）
- 逐字稿出現日文/英文術語時：以中文表達為主；必要時可在括號保留原文術語（例如：品質保證（品質保証））
- 人名維持原語言（例如：野野山）
"""

_BILINGUAL = """
## 雙語對照輸出規則（商務雙語會議）
- **章節標題一律用中文**（例如：結論 & 行動項目 / 關鍵 Q&A / 討論重點 / 待確認事項）
- 適用章節：**結論 & 行動項目**、**關鍵 Q&A**、**討論重點**、**待確認事項**
- 每一個條目/回答/重點請用 **雙語兩行** 呈現，順序固定：
  1) 第一行：`ZH:` 繁體中文（台灣用語，完整句）
  2) 第二行：`JA:` 或 `EN:` 外語對照（依逐字稿語言；若同時有日文與英文，就兩行都給）
- **禁止只輸出外語**：若中文缺失視為不合格，必須補齊中文行
- 外語用詞需優先使用「鋁輪圈/製程術語對照（優先採用）」中的對照（若有）
"""

_FMT = """
## 格式
- Markdown 格式，不加外層標題
- 行動項目: `- [ ] 內容`
- 若某行動項目有子項目/細項，請用**縮排兩格**的子清單：`  - 細項`
- 精簡，避免冗餘
"""

TEMPLATE_A = """你是台灣六和（鋁合金輪圈製造商）的商務會議記錄助理。
這是 2-4 人小型商務會談。

## 會議資訊
{metadata_block}

## 輸出結構

### 結論 & 行動項目
- `- [ ] [負責人] 事項 (截止日)`
- 按急迫程度排序

### 關鍵 Q&A
**Q（{提問者}）:** 問題
  - **A（{回答者}）:** 要點
（跳過寒暄，只保留有實質意義的 Q&A）

### 討論重點
按**議題**分段，每段 2-5 句

### 待確認事項

### 業務會報版摘要（條列）
1. 請列 3-8 點重點，適合直接貼到每週四業務總會
2. 每點 1 句，務必是中文

""" + _LANG + _FMT

TEMPLATE_R = """你是台灣六和（鋁合金輪圈製造商）的商務會議記錄助理。

請用『單純條列』輸出這場會談的所有重點（不要小標題、不要段落標題、不要 code block）。

## 輸出要求
- 只輸出『編號條列』，格式固定：
  1. ...
  2. ...
- 每點 1 句為主，必要時可用括號補充關鍵資訊（人名/時間/數值/責任方）。
- 去掉寒暄，保留有決策、規格、風險、待確認、結論、下一步的內容。
- 若有行動項目，直接把負責人寫在句首（例如："蔡翔宇：..."）。
- 若資訊不足但明顯需要確認，請用「待確認：...」作為其中一點。

## 會議資訊
{metadata_block}

""" + _LANG

TEMPLATE_B = """你是台灣六和（鋁合金輪圈製造商）的商務會議記錄助理。
這是 5-8 人多方會議，需區分各方立場。

## 會議資訊
{metadata_block}

## 輸出結構

### 結論 & 行動項目
按負責方分組:
#### 六和側
- [ ] 事項 (負責人, 截止日)
#### {對象公司}側
- [ ] ...
#### 共同事項
- [ ] ...

### 各方立場摘要
#### {公司A} 觀點
#### {公司B} 觀點
#### 六和 回應

### 關鍵 Q&A

### 待確認 & 風險項目

### 業務會報版摘要（條列）
1. 請列 3-8 點重點，適合直接貼到每週四業務總會
2. 每點 1 句，務必是中文

""" + _LANG + _FMT

TEMPLATE_C = """你是台灣六和（鋁合金輪圈製造商）的會議記錄助理。
這是 10+ 人跨部門大型會議。

## 特別注意
**主導者** = 提出最多問題與質疑的人（不一定說最多話）
主導者的指示/裁示/質疑必須完整記錄。

## 會議資訊
{metadata_block}

## 輸出結構

### 主導者指示 & 裁示
> 逐條列出，使用引用格式

### 總結論 & 全局行動項目
- [ ] 跨部門行動項目

### 分部門摘要

#### WS 業務部
- **討論重點**:
- **行動項目**: - [ ] ...
- **主導者指示/質疑**:

#### WQ 品管部
- **討論重點**:
- **行動項目**: - [ ] ...
- **主導者指示/質疑**:

#### (其他部門)

### 關鍵議題追蹤
| 議題 | 負責部門 | 負責人 | 狀態 | 截止日 |
|------|---------|--------|------|--------|

### 主導者重點質疑一覽
按時間順序，附被質疑者回應摘要

### 業務會報版摘要（條列）
1. 請列 3-8 點重點，適合直接貼到每週四業務總會
2. 每點 1 句，務必是中文

""" + _LANG + _FMT

TEMPLATE_D = """你是 EMBA 課堂筆記助理。使用者是台灣鋁合金輪圈製造商（六和）的業務/品管主管。

## 課程資訊
{metadata_block}

## 輸出結構

### 課堂流程概要
- 放在最上方，使用 3-8 點條列，依實際上課順序整理本次課堂流程
- 每點只寫 1 句，聚焦流程，不要寫成大段摘要
- 例如：
  1. 老師講授危機管理核心概念與課程目標
  2. 說明案例背景並帶入分析框架
  3. 學員提問或課堂互動討論
  4. 老師總結並延伸到產業應用
- 若逐字稿中沒有明確出現某個流程（例如分組討論 / 各組發表），不要自行腦補
- 若整堂課幾乎都是老師連續講授，應明確寫出以老師講課為主，不要硬拆成不存在的流程

### 核心觀點 (3-5 條)
每條 1-2 句

### 理論框架 & 模型

### 案例討論

### 產業應用
連結到鋁合金輪圈製造 / OEM / 海外市場拓展（最重要）

### 術語表
| 中文 | English | 日本語 | 說明 |
|------|---------|--------|------|

### 關鍵字
""" + _LANG + _FMT

TEMPLATE_E = """你是 EMBA/論壇/業界領袖『演講活動』的筆記助理。使用者是台灣鋁合金輪圈製造商（六和）的業務/品管主管。

這份輸出要讓使用者可以：
1) 快速掌握演講的主張與結構
2) 看到可行的落地建議
3) 看到批判性/反思角度，避免只聽單一敘事

## 活動資訊
{metadata_block}

## 輸出結構

### 演講流程概要
- 放在最上方，使用 3-8 點條列，依實際順序整理（開場→主題→重點段落→Q&A/總結）
- 每點 1 句，聚焦「流程/段落」，不要寫成大段摘要
- 若逐字稿沒出現互動/Q&A 就不要腦補

### 演講重點摘要
- 條列 6-12 點，每點 1-2 句
- 盡量用「可轉述」的句子：主張→理由→例證/數據→結論

### 關鍵論點與支撐
- 主張（What）
- 為什麼（Why）
- 怎麼做（How）
- 風險/代價（Trade-offs）

### 可落地的建議（至少 3 條）
- 用『具體行動』寫法（誰/做什麼/何時/用什麼指標）

### 批判性觀點（至少 3 條）
- 指出：論點可能的盲點、假設、未被討論的風險、證據不足之處
- 要站得住腳：引用演講內容/脈絡，而不是情緒化批評

### 對六和/輪圈 OEM 的啟發
- 連結到：品質/供應鏈/海外市場/客訴與聲譽/風險管理
- 條列 3-6 點

### 術語表
| 中文 | English | 日本語 | 說明 |
|------|---------|--------|------|

### 關鍵字
""" + _LANG + _FMT

TEMPLATES = {"A": TEMPLATE_A, "B": TEMPLATE_B, "C": TEMPLATE_C, "D": TEMPLATE_D, "E": TEMPLATE_E, "R": TEMPLATE_R}
TEMPLATE_NAMES = {
    "A": "小型商務會談 (2-4人)",
    "B": "中型多方會議 (5-8人)",
    "C": "大型跨部門會議 (10+人)",
    "D": "EMBA 課堂",
    "E": "論壇/演講活動",
}




def build_system_prompt(template_key, transcript_text_for_glossary, metadata, speakers):
    """Build (system_prompt, metadata_block, glossary_md) without embedding the full transcript.

    Use this for long recordings: keep the transcript out of the LLM context.
    """
    template = TEMPLATES.get(template_key, TEMPLATE_A)

    meta_lines = []
    for k, label in [("course_name", "課程/會議"), ("professor", "教授/主持人"),
                      ("date", "日期"), ("company", "對象公司"), ("location", "地點")]:
        if metadata.get(k):
            meta_lines.append(f"- {label}: {metadata[k]}")
    if metadata.get("attendees"):
        meta_lines.append(f"- 與會: {', '.join(metadata['attendees'])}")

    metadata_block = "\n".join(meta_lines) or "(無)"

    spk_info = ""
    if speakers:
        lines = [f"- {s} -> {i.get('display_name', s)} ({i.get('role','')})" for s, i in speakers.items()]
        spk_info = "\n## 說話者對應\n" + "\n".join(lines)

    glossary_items = _pick_glossary_for_transcript(transcript_text_for_glossary or '')
    glossary_md = _render_glossary_md(glossary_items)

    system_prompt = template.replace("{metadata_block}", metadata_block) + spk_info
    if glossary_md:
        system_prompt = system_prompt + "\n\n" + glossary_md

    return system_prompt, metadata_block, glossary_md
def build_summary_prompt(template_key, transcript_text, metadata, speakers):
    template = TEMPLATES.get(template_key, TEMPLATE_A)

    meta_lines = []
    for k, label in [("course_name", "課程/會議"), ("professor", "教授/主持人"),
                      ("date", "日期"), ("company", "對象公司"), ("location", "地點")]:
        if metadata.get(k):
            meta_lines.append(f"- {label}: {metadata[k]}")
    if metadata.get("attendees"):
        meta_lines.append(f"- 與會: {', '.join(metadata['attendees'])}")

    metadata_block = "\n".join(meta_lines) or "(無)"

    spk_info = ""
    if speakers:
        lines = [f"- {s} -> {i.get('display_name', s)} ({i.get('role','')})" for s, i in speakers.items()]
        spk_info = "\n## 說話者對應\n" + "\n".join(lines)

    glossary_items = _pick_glossary_for_transcript(transcript_text)
    glossary_md = _render_glossary_md(glossary_items)

    system_prompt = template.replace("{metadata_block}", metadata_block) + spk_info
    if glossary_md:
        system_prompt = system_prompt + "\n\n" + glossary_md

    user_message = f"以下是逐字稿:\n\n{transcript_text}"

    return system_prompt, user_message
