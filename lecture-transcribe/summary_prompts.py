from __future__ import annotations

"""
summary_prompts.py — LLM 摘要 Prompt 模板
A: 小型商務 (2-4人) | B: 中型多方 (5-8人) | C: 大型跨部門 (10+人) | D: EMBA 課堂
W: WS 業務週會（含各客戶待辦 & 主管指示）
"""

import re
from pathlib import Path
from typing import List, Dict, Tuple

_REFERENCE_DIR = Path(__file__).parent / 'references'
_FEW_SHOT_FILES = {
    'D4': [
        _REFERENCE_DIR / 'few-shot-d4-mixed.md',
        _REFERENCE_DIR / 'few-shot-d4-crossculture-lecture-heavy.md',
    ],
}

# ============================================================
# Glossary: 鋁輪圈/製程 中日英術語對照
# - 來源：memory/鋁輪圈專有名詞.xlsx（由使用者提供）
# - 用途：只把『逐字稿中有出現的術語』挑出來塞進 prompt，避免 prompt 爆長
# ============================================================

_CANDIDATE_GLOSSARY_PATHS = [
    Path.home() / '.openclaw' / 'workspace' / 'memory' / '鋁輪圈專有名詞.xlsx',
    Path.home() / '.openclaw' / 'workspace' / 'memory' / 'wheel_glossary.xlsx',
]


def _load_wheel_glossary_rows(xlsx_path: Path) -> List[Dict[str, str]]:
    '''Load glossary rows from an xlsx file.

    Columns are flexible; we try to match headers containing:
      - 中文 / 英文 / 日文 / 說明(備註)
    '''
    try:
        from openpyxl import load_workbook
    except Exception:
        return []

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        return []

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(c).strip() if c is not None else '' for c in rows[0]]

    def find_col(keys):
        for k in keys:
            for i, h in enumerate(header):
                if k in h:
                    return i
        return None

    zh_i = find_col(['中文', '中', '繁'])
    en_i = find_col(['英文', '英', 'EN'])
    jp_i = find_col(['日文', '日', 'JP', '日本語'])
    note_i = find_col(['說明', '備註', '註', 'note'])

    out = []
    for r in rows[1:]:
        def get(i):
            if i is None or i >= len(r):
                return ''
            v = r[i]
            return str(v).strip() if v is not None else ''

        item = {
            'zh': get(zh_i),
            'en': get(en_i),
            'jp': get(jp_i),
            'note': get(note_i),
        }
        if any(item.values()):
            out.append(item)
    return out


def _pick_glossary_for_transcript(transcript_text: str, max_items: int = 60) -> List[Dict[str, str]]:
    xlsx = next((pp for pp in _CANDIDATE_GLOSSARY_PATHS if pp.exists()), None)
    if not xlsx:
        return []

    rows = _load_wheel_glossary_rows(xlsx)
    if not rows:
        return []

    t = transcript_text or ''
    picked = []
    for it in rows:
        keys = [it.get('jp', ''), it.get('en', ''), it.get('zh', '')]
        keys = [k for k in keys if k and len(k) >= 2]
        if keys and any(k in t for k in keys):
            picked.append(it)
            if len(picked) >= max_items:
                break
    return picked


def _render_glossary_md(items: List[Dict[str, str]]) -> str:
    if not items:
        return ''
    lines = [
        '## 鋁輪圈/製程術語對照（優先採用）',
        '- 若逐字稿出現下列術語，請優先使用這份對照做翻譯/統一用詞',
        '- 若沒有對照，保留原文並在括號補繁中即可（不要自造名詞）',
        '',
        '| 中文 | English | 日本語 | 說明 |',
        '|---|---|---|---|',
    ]
    for it in items:
        lines.append(f"| {it.get('zh','')} | {it.get('en','')} | {it.get('jp','')} | {it.get('note','')} |")
    return "\n".join(lines) + "\n\n"



_EMBA_TEMPLATE_KEYS = ("D1", "D2", "D3", "D4")


def _route_text_blob(transcript_text: str = '', metadata: dict | None = None) -> str:
    metadata = metadata or {}
    parts = [
        transcript_text or '',
        metadata.get('course_name') or '',
        metadata.get('meeting_name') or '',
        metadata.get('topic') or '',
        metadata.get('reference_notes') or '',
        ' '.join(metadata.get('keywords') or []),
    ]
    return "\n".join([p for p in parts if p]).lower()


def is_global_taiwan_case_course(metadata: dict | None = None, transcript_text: str = '') -> bool:
    metadata = metadata or {}
    text = _route_text_blob(transcript_text, metadata)
    selected = (metadata.get('selected_template') or '').strip().upper()
    override_u = (metadata.get('template_override') or '').strip().upper()
    if '全球台商個案' in text:
        return True
    if selected == 'D2' or override_u == 'D2':
        course_name = (metadata.get('course_name') or '').strip()
        if '個案' in course_name or '台商' in course_name:
            return True
    return bool(metadata.get('global_taiwan_case_mode'))


def apply_course_specific_summary_flags(metadata: dict | None = None, transcript_text: str = '') -> dict:
    metadata = metadata or {}
    if is_global_taiwan_case_course(metadata, transcript_text):
        metadata['global_taiwan_case_mode'] = True
        metadata['detailed_case_restore'] = True
        metadata['enable_company_web_enrichment'] = True
    return metadata


def _add_hit(hits: dict, key: str, label: str, limit: int = 10) -> None:
    if label not in hits[key] and len(hits[key]) < limit:
        hits[key].append(label)


def _count_any(text: str, keywords: list[str], per_kw_cap: int = 3) -> tuple[int, list[str]]:
    score = 0
    hit_labels = []
    for kw in keywords:
        c = text.count(kw.lower())
        if not c:
            continue
        weight = 3 if len(kw) >= 3 else 1
        score += min(c, per_kw_cap) * weight
        hit_labels.append(kw)
    return score, hit_labels


def detect_emba_subtemplate(transcript_text: str = '', metadata: dict | None = None) -> Tuple[str, dict]:
    metadata = metadata or {}
    text = _route_text_blob(transcript_text, metadata)

    keyword_map = {
        'D1': [
            '理論', '模型', '架構', '研究', '變項', '假設', '文獻', '定義', '概念', '命題',
            '量表', '問卷', '研究方法', '信度', '效度', '推論', '分析', '方法論', '機制', '演變',
            '調節', '中介', '依變項', '自變項', '構面', '內涵',
        ],
        'D2': [
            '個案', 'case', '策略', '決策', '風險', '競爭', '市場', '定價', '選擇', '衝突',
            '轉型', '失敗', '成功', '情境', '討論', '分組討論', '如果是你', 'swot', '五力', 'stp',
            '贊成', '反對', '該不該', '值不值得', '怎麼判斷', '取捨', '權衡',
        ],
        'D3': [
            '集合', '出發', '航班', '機場', '住宿', '費用', '匯款', '護照', '簽證', '行李',
            '托運', '手提', '參訪', '工作坊', '自由活動', 'grab', 'esim', '飯店', '機票', '回程',
            '翻譯', '接送', '集合時間', '注意事項', '提醒', '報到', '護照效期', '電子簽證',
            '截止', '自備', '穿著', '集合地點', '候補', '候選', '機場集合', '審查', '備案',
            '改期', '取消', '行程', '交流', '自由活動',
        ],
        'D4': [
            '第一組', '第二組', '第三組', '第1組', '第2組', '第3組', '小組', '分組', '簡報', '報告',
            '講評', 'q&a', 'q & a', '討論題', '討論問題', '如果我是學生', '老師提醒', '老師補充',
            '交流', '參訪', '行程', '注意事項', '備案', '取消', '工作坊',
        ],
    }

    scores = {k: 0 for k in _EMBA_TEMPLATE_KEYS}
    hits = {k: [] for k in _EMBA_TEMPLATE_KEYS}

    for key, kws in keyword_map.items():
        sc, labels = _count_any(text, kws)
        scores[key] += sc
        for label in labels:
            _add_hit(hits, key, label)

    if re.search(r'\bjx\d{3,4}\b|\btpe\b|\bsgn\b|\b\d{1,2}:\d{2}\b', text):
        scores['D3'] += 8
        _add_hit(hits, 'D3', '航班/時間格式')
    if re.search(r'\b\d{1,2}/\d{1,2}\b|\d{4}-\d{2}-\d{2}|週[一二三四五六日天]', text):
        scores['D3'] += 3
        _add_hit(hits, 'D3', '日期/週期格式')
    if re.search(r'twd|usd|nt\$|\$|元/人|kg/人|每人|兩人一室', text):
        scores['D3'] += 4
        _add_hit(hits, 'D3', '費用/規格格式')
    if re.search(r'checklist|待辦|截止|deadline|先做|盡快|務必|記得', text):
        scores['D3'] += 4
        _add_hit(hits, 'D3', '待辦/提醒語氣')

    if re.search(r'信度|效度|研究設計|依變項|自變項|中介|調節|問卷|量表', text):
        scores['D1'] += 8
        _add_hit(hits, 'D1', '研究設計/變項語彙')
    if re.search(r'定義|概念|內涵|框架|模型|推論|命題|理論', text):
        scores['D1'] += 4
        _add_hit(hits, 'D1', '理論拆解語氣')
    if re.search(r'差異|比較|演變|階段|構面|機制', text):
        scores['D1'] += 3
        _add_hit(hits, 'D1', '比較/演變語氣')

    if re.search(r'如果是你|你會怎麼|該不該|是否應該|贊成還是反對|要不要', text):
        scores['D2'] += 8
        _add_hit(hits, 'D2', '案例判斷語氣')
    if re.search(r'個案公司|案例公司|案例|情境|困境|抉擇|利弊|取捨|權衡', text):
        scores['D2'] += 5
        _add_hit(hits, 'D2', '案例情境語氣')
    q_count = text.count('？') + text.count('?')
    if q_count >= 3:
        scores['D2'] += min(q_count, 6)
        _add_hit(hits, 'D2', '高問答密度')

    presentation_score = 0
    admin_score = 0
    if re.search(r'第[一二三四五六七八九十123456789]組|group\s*[123456789]', text):
        presentation_score += 8
        _add_hit(hits, 'D4', '多組報告結構')
    if re.search(r'講評|comment|feedback|q\s*&\s*a|q&a|討論題|討論問題|如果我是學生', text):
        presentation_score += 8
        _add_hit(hits, 'D4', '老師講評 / Q&A')
    if re.search(r'簡報|報告|presentation', text):
        presentation_score += 4
        _add_hit(hits, 'D4', '報告/簡報語境')
    if re.search(r'行程|參訪|工作坊|交流|審查|取消|備案|注意事項|自由活動', text):
        admin_score += 8
        _add_hit(hits, 'D4', '行政/行程/備案')
    if re.search(r'航班|機場|住宿|grab|冰塊|路邊攤|腸胃藥', text):
        admin_score += 6
        _add_hit(hits, 'D4', '在地提醒/實務準備')
    if re.search(r'待辦|deadline|截止|盡快|務必|先做', text):
        admin_score += 4
        _add_hit(hits, 'D4', '後續行動')
    if presentation_score >= 8 and admin_score >= 8:
        scores['D4'] += presentation_score + admin_score + 12
        _add_hit(hits, 'D4', '講評 + 行政混搭')
    else:
        scores['D4'] += max(presentation_score, admin_score)

    course_name = (metadata.get('course_name') or '').lower()
    if any(k in course_name for k in ('研究方法', '消費者行為', '方法論')):
        scores['D1'] += 5
        _add_hit(hits, 'D1', '課名先驗')
    if any(k in course_name for k in ('個案', '危機管理', '策略')):
        scores['D2'] += 5
        _add_hit(hits, 'D2', '課名先驗')
    if any(k in course_name for k in ('跨文化交流', '研習', '參訪')):
        scores['D3'] += 5
        _add_hit(hits, 'D3', '課名先驗')
        if presentation_score >= 8:
            scores['D4'] += 5
            _add_hit(hits, 'D4', '課名先驗')

    priority = {'D4': 4, 'D3': 3, 'D2': 2, 'D1': 1}
    ranked = sorted(_EMBA_TEMPLATE_KEYS, key=lambda k: (scores[k], priority[k]), reverse=True)
    best = ranked[0] if scores[ranked[0]] > 0 else 'D1'
    reason_hits = '、'.join(hits[best][:5]) if hits[best] else '未命中特定高權重線索，先走理論/講授版'
    reason = f"自動路由 → {best}（命中：{reason_hits}）"
    return best, {'scores': scores, 'hits': hits, 'reason': reason}


def _resolve_emba_template_key(metadata: dict, transcript_text: str = '') -> str:
    selected = (metadata.get('selected_template') or '').strip().upper()
    if selected in _EMBA_TEMPLATE_KEYS:
        return selected
    override_u = (metadata.get('template_override') or '').strip().upper()
    if override_u == 'D':
        return 'D1'
    if override_u in _EMBA_TEMPLATE_KEYS:
        return override_u
    detected, _ = detect_emba_subtemplate(transcript_text, metadata)
    return detected


def _emba_chunk_focus(template_key: str) -> tuple[str, str]:
    if template_key == 'D1':
        return (
            '這一段是 EMBA 理論 / 教授講授型內容。優先抓：核心概念、老師如何拆解、容易混淆處、可直接拿去寫作業/發言的句子。另外也要記錄這一段的「教學活動流程」（老師做了什麼、用了什麼方式講、有沒有舉例/板書/影片），供後續合併成完整課堂流程還原。',
            '''### 本段關鍵觀念
- ...

### 老師怎麼拆這個概念
- ...

### 容易混淆 / 容易考錯
- ...

### 可直接拿去寫作業 / 發言
- ...

### 本段教學活動流程
- 老師在這段做了什麼（講授/舉例/板書/影片/討論/點名）、大致順序、用了什麼案例或素材

### 名詞/術語
- 術語：中文（原文） — 1 句解釋''',
        )
    if template_key == 'D2':
        return (
            '這一段是 EMBA 個案 / 討論判斷型內容。若提到多家公司，必須逐案拆開，不可混成抽象總結。優先抓：每個 case 的公司背景、所在市場、商業模式、核心衝突、老師的判斷依據、可直接拿去報告/發言的論點。另外也要記錄這一段的「教學活動流程」（案例怎麼帶入、討論怎麼進行、老師怎麼收束），供後續合併成完整課堂流程還原。',
            '''### 本段台商個案逐案還原
- 公司／案例 1：
  - 公司背景 / 所在市場 / 商業模式
  - 核心衝突 / 決策難題
  - 老師怎麼看
  - 可直接拿去發言的論點

### 本段 case 背景 / 情境
- ...

### 核心衝突 / 真正問題
- ...

### 老師的判斷 / 分析依據
- ...

### 可直接拿去報告 / 發言的論點
- ...

### 本段教學活動流程
- 老師在這段做了什麼（案例引入/提問/分組討論/點評/影片/收束）、大致順序

### 待確認 / 保留意見
- ...''',
        )
    if template_key == 'D4':
        return (
            '這一段是 EMBA「報告講評 + 行政/行程說明」混搭內容。優先抓：老師對各組的核心建議、指正/補充、共同提醒，以及已確認/可能變動的行程與注意事項。對老師講評請低抽象整理、儘量還原原意，不要替老師延伸加料。',
            '''### 本段老師講評 / 指正
- 依組別整理：第一組 / 第二組 / 第三組
- 每組只收斂成：核心建議 / 指正或補充

### 本段共同提醒 / Q&A 焦點
- ...

### 本段行政 / 行程 / 備案
- 已確認安排 / 可能變動 / 備案

### 本段注意事項 / 待辦
- ...

### 本段說明活動流程
- 這段先講了哪一組、老師怎麼講評、何時切到行政/行程說明、有沒有 QA
''',
        )
    return (
        '這一段是 EMBA 行政 / 參訪 / 任務說明型內容。優先抓：已確認事項、時間/地點/費用/任務、自備清單、提醒與下一步。另外也要記錄這一段的「說明活動流程」（先講什麼後講什麼、QA 環節問了什麼），供後續合併成完整課堂流程還原。',
        '''### 本段已確認事項
- ...

### 行程 / 任務 / 時間點
- ...

### 最容易忘但很重要
- ...

### 行動項目
- [ ] ...

### 本段說明活動流程
- 這段先講了什麼、後講了什麼、有沒有 QA、誰問了什麼

### 可能變動 / 待確認
- ...''',
    )


def select_template(meeting_type, speaker_count, duration_min, override=None, transcript_text='', metadata=None):
    override_u = (override or '').upper().strip()
    if meeting_type == "emba":
        if isinstance(metadata, dict):
            apply_course_specific_summary_flags(metadata, transcript_text)
        if override_u == 'D':
            override_u = 'D1'
        if override_u in ("D1", "D2", "D3", "D4", "E"):
            if isinstance(metadata, dict):
                metadata['selected_template_reason'] = f'手動指定模板 → {override_u}'
            return override_u
        detected, route_info = detect_emba_subtemplate(transcript_text, metadata if isinstance(metadata, dict) else {})
        if isinstance(metadata, dict):
            metadata['selected_template_reason'] = route_info.get('reason')
            metadata['selected_template_scores'] = route_info.get('scores')
            metadata['selected_template_hits'] = route_info.get('hits')
        return detected

    if override_u and override_u in ("A", "B", "C", "E", "R", "W"):
        return override_u
    if speaker_count > 8 or duration_min > 120:
        return "C"
    elif speaker_count >= 5 or duration_min >= 60:
        return "B"
    return "A"


def select_model(duration_min, user_pref=None):
    m = {"codex": "openai-codex/gpt-5.4",
         "gpt": "openai-codex/gpt-5.4",
         "gemini": "google/gemini-3.1-pro-preview"}
    if user_pref and user_pref.lower() in m:
        return m[user_pref.lower()]
    # 預設走 openai-codex OAuth（包月）
    return m["codex"]


from typing import Optional

def select_chunk_model(duration_min: float, user_pref: Optional[str] = None) -> str:
    """Pick a high C/P model for per-chunk processing.

    Goal: speed + low cost + low rate-limit risk.

    Chunk stage優先便宜快，預設走使用者已包月的 MiniMax M2.7；
    若失敗，再由 lecture_pipeline.py 的 chunk fallback chain 接手。
    """
    return "minimax-portal/MiniMax-M2.7"


def _render_reference_entities(metadata: dict) -> str:
    items = metadata.get('reference_entities') or []
    if not items:
        return ''
    lines = [
        '## 使用者提供的名稱/名詞校對清單（高優先）',
        '- 用途：優先用於校正人名、公司名、講者名稱、活動名稱與專有名詞，避免逐字稿同音錯字。',
        '- 原則：若逐字稿出現近音、近形或疑似 ASR 錯字，優先採用此清單的正確寫法；但不要憑空新增逐字稿完全未提及的內容。',
        '',
    ]
    lines.extend([f'- {x}' for x in items[:32]])
    return "\n".join(lines) + "\n"


def _render_reference_notes(metadata: dict) -> str:
    notes = (metadata.get('reference_notes') or '').strip()
    if not notes:
        return ''
    detail_flag = '是' if metadata.get('detailed_material_restore') else '否'
    exam_flag = '是' if metadata.get('high_priority_exam_material') else '否'
    skip_notion_ai_flag = '是' if metadata.get('skip_notion_ai_transcript') else '否'
    notion_ai_fallback_flag = '是' if metadata.get('use_notion_ai_transcript_fallback') else '否'
    original_quality = metadata.get('original_transcript_quality') or 'unknown'
    return (
        "\n\n## 使用者提供的參考資料（高優先參考，但不可凌駕逐字稿事實）\n"
        "- 可能包含：使用者筆記、Notion 筆記 / AI 摘要 / AI transcript、圖片 OCR、教材附件節錄、前段缺錄補充。\n"
        f"- 本次教材詳細還原模式：detailed_material_restore={detail_flag}；high_priority_exam_material={exam_flag}。\n"
        f"- 本次 Notion AI 逐字稿策略：skip_notion_ai_transcript={skip_notion_ai_flag}；fallback_used={notion_ai_fallback_flag}；original_transcript_quality={original_quality}。\n"
        "- 用途：幫助交叉核對重點、修正術語/人名、補強結構與心得寫作角度，並優化摘要完整性。\n"
        "- ⚠️ 特別注意：使用者筆記中可能記錄了『錄音無法捕捉的內容』，例如課堂影片摘要、板書、老師口頭補充、作業/報告指示。\n"
        "  這些資訊不在逐字稿中，但同等重要，必須納入對應章節，不可因逐字稿沒有就忽略。\n"
        "- ⚠️ 若參考資料屬於教材翻拍照片 / 投影片 / 講義 OCR，請把它視為『教材內容還原素材』，不是普通補充筆記。\n"
        "  遇到定義、分類、構面、步驟、比較表、判準、條列、考點時，必須盡量具體還原；寧可保留較多細節，也不要壓成 4-6 個空泛 bullets。\n"
        "- 請保留來源類別，不要把不同來源混寫成同一種證據。至少要分清：使用者筆記補充、教材照片 OCR / 圖片 OCR 補充、前段缺錄補足、Notion AI transcript / summary。\n"
        "- 若 skip_notion_ai_transcript=true，Notion AI 長逐字稿預設只當 reference，不可覆蓋錄音原始轉錄。\n"
        "- 只有當 fallback_used=true，或錄音原始轉錄明顯語意斷裂時，才可保守參考 Notion AI 逐字稿補足上下文；若兩者衝突，優先採信逐字稿與可互證內容。\n"
        "- 原則：若參考資料與逐字稿衝突，以逐字稿為準；若是逐字稿沒有但筆記有的補充內容，請依章節規則納入。\n\n"
        f"{notes}\n"
    )


def _render_external_company_enrichment(metadata: dict) -> str:
    items = metadata.get('company_web_enrichment') or []
    if not items:
        return ''

    lines = [
        '## 企業背景補充（外部資訊，不是課堂原話）',
        '- 以下內容來自外部查詢 / web search enrichment，只能當背景補充，不能偽裝成老師原話或課堂逐字稿。',
        '- 用途：補足公司背景、產業位置、近期發展、風險與為何值得被當作個案。',
        '- 寫入最終摘要時，請明確標示這是「外部補充資訊」。',
        '',
    ]
    for idx, item in enumerate(items[:5], 1):
        if not isinstance(item, dict):
            continue
        company = (item.get('company') or item.get('name') or f'公司 {idx}').strip()
        query = (item.get('query') or '').strip()
        summary = (item.get('summary') or item.get('background') or '').strip()
        raw = (item.get('raw') or '').strip()
        lines.append(f'### {idx}. {company}')
        if query:
            lines.append(f'- 查詢：{query}')
        if summary:
            lines.append(summary)
        elif raw:
            lines.append(raw)
        lines.append('')
    return "\n".join(lines).strip() + "\n"


def _render_case_mode_rules(metadata: dict, template_key: str | None = None) -> str:
    if template_key != 'D2' and not metadata.get('global_taiwan_case_mode'):
        return ''

    if not (metadata.get('detailed_case_restore') or metadata.get('global_taiwan_case_mode')):
        return ''

    return (
        '## 全球台商個案 / D2 專屬規則\n'
        '- 這堂課是個案課，重點不是抽象結論，而是逐案還原。\n'
        '- 只要 transcript / 使用者筆記 / OCR / 教材有提到具體公司，就要盡量以「一家公司一家公司」方式整理，不要混成一句「老師舉了幾個例子」。\n'
        '- 最終摘要中，請穩定產出 `### 台商個案逐案還原` 章節。\n'
        '- 每個個案優先保留：公司背景、所在市場、商業模式、核心衝突、老師判斷、可拿去發言的論點。\n'
        '- 若同時提到多家公司，要分開整理；不要把不同公司的情境、風險、決策混在同一段。\n'
        '- 若有外部企業背景補充，請另外放在 `### 企業背景補充（外部資訊）`，並清楚標示不是老師原話。\n'
    )


# ============================================================
# User-priority item extractor
# 從 reference_notes 抽出「不在錄音主線但使用者非常需要知道的高價值資訊」
# 關鍵字類型：作業/報告/考試/行政、影片補充、老師額外提醒
# ============================================================

_USER_PRIORITY_KEYWORD_GROUPS = {
    'assignment': [
        '期中', '期末', '作業', '報告', '考試', '截止', 'deadline',
        'ppt', 'powerpoint', '簡報', '每人', '每人分鐘', '字數', '格式', '加分',
        '要交', '繳交', '評分', '佔分', '分組報告',
    ],
    'video': [
        '影片', '影片鑑賞', '影片補充', '案例分享', '影片案例', '課堂影片',
        '老師補充', '額外補充', '延伸補充', '老師說', '老師提到',
    ],
    'teacher_note': [
        '老師金句', '老師提醒', '老師強調', '重要提醒', '注意', '特別提醒',
        '辨別', '要記', '千萬', '務必', '一定要',
    ],
}

def extract_user_priority_items(reference_notes: str) -> dict[str, list[str]]:
    """從 reference_notes 抽出 user-priority 類型的條目。

    Returns:
        dict with keys: 'assignment', 'video', 'teacher_note'
        each value is a list of matching lines
    """
    if not reference_notes:
        return {}

    result: dict[str, list[str]] = {}
    lines = reference_notes.splitlines()

    for line in lines:
        line_stripped = line.strip()
        if not line_stripped:
            continue
        line_lower = line_stripped.lower()
        for group, keywords in _USER_PRIORITY_KEYWORD_GROUPS.items():
            if any(kw in line_lower for kw in keywords):
                result.setdefault(group, [])
                if line_stripped not in result[group]:
                    result[group].append(line_stripped)
                break  # each line goes to first matching group

    return result


def render_user_priority_supplement(reference_notes: str, metadata: dict | None = None) -> str:
    """為 EMBA 摘要生成『使用者筆記補充』章節的 prompt 指令。

    若有抽到任何 user-priority items，回傳要插入 system prompt 的強制指令區塊。
    若沒有，回傳空字串。
    """
    items = extract_user_priority_items(reference_notes)
    if not items:
        return ''

    sections = []

    assignment_items = items.get('assignment', [])
    if assignment_items:
        bullet_lines = '\n'.join(f'  - {x}' for x in assignment_items)
        sections.append(
            f"### 📋 作業 / 報告 / 行政提醒\n"
            f"（以下來自使用者筆記，不在錄音主線，但必須完整呈現）\n"
            f"{bullet_lines}"
        )

    video_items = items.get('video', [])
    if video_items:
        bullet_lines = '\n'.join(f'  - {x}' for x in video_items)
        sections.append(
            f"### 🎬 課堂影片 / 課外補充重點\n"
            f"（錄音無法完整捕捉，以下來自使用者筆記，請完整保留）\n"
            f"{bullet_lines}"
        )

    teacher_notes = items.get('teacher_note', [])
    if teacher_notes:
        bullet_lines = '\n'.join(f'  - {x}' for x in teacher_notes)
        sections.append(
            f"### 💬 老師額外提醒 / 金句補充\n"
            f"（以下來自使用者筆記記錄，請納入對應章節）\n"
            f"{bullet_lines}"
        )

    if not sections:
        return ''

    header = (
        "\n\n## ⚠️ 使用者筆記高優先補充（必須納入最終摘要，不可省略）\n"
        "以下內容來自使用者自己的 Notion 筆記，可能是課堂影片摘要、作業指示、老師補充，\n"
        "這些資訊不在錄音或逐字稿中，但對使用者同樣重要甚至更重要。\n"
        "⚠️ 規則：以下每一個類別，若有內容，最終摘要中必須出現對應章節，不能因為逐字稿沒有就跳過。\n\n"
    )
    return header + '\n\n'.join(sections) + '\n'


def build_chunk_prompt(
    transcript_chunk: str,
    chunk_label: str,
    metadata: dict,
    speakers: dict,
    bilingual: bool = False,
):
    """Return (system_prompt, user_message) for chunk summarization.

    Output must be compact and merge-friendly.
    """

    if metadata.get('type') == 'emba':
        template_key = _resolve_emba_template_key(metadata, transcript_chunk)
        focus_rule, output_schema = _emba_chunk_focus(template_key)
        sys = f"""你是 EMBA 課堂逐字稿的『分段整理器』。

目標：把這一小段逐字稿整理成後續可合併的高密度筆記，避免冗詞。
規則：
- 只輸出繁體中文（台灣用語），除非必要術語需要括號保留原文。
- 不要重述逐字稿；只抓資訊密度高、之後真的會回頭看的內容。
- 口吻要像整理給自己看的筆記：直白、實用、有判斷感，但不能超出逐字稿證據。
- 避免 AI 套話 / 公文腔 / 翻譯腔，例如「本段主要在講…」「此外」「首先／其次／最後」「綜上所述」；直接寫重點。
- 若逐字稿不是在引用原話，優先改成台灣常用說法；像「數據／質量／支持／落地／跟進」這類詞，請依語境改寫成「資料／品質／支援／落實／追蹤」等自然用法。
- 若參考資料中有教材照片 OCR / 投影片 / 講義內容，請同步抽出這一段對應的教材資訊，尤其是定義、分類、構面、步驟、比較點、判準、考點，不要只寫成一句泛化摘要。
- 目前子模板：{template_key}
- 聚焦方向：{focus_rule}
- 只輸出下列章節，缺就省略；不要硬湊。

輸出格式（Markdown）：
### 分段資訊
- 段落：{chunk_label}

{output_schema}
"""
    else:
        # Build client-person mapping block if available
        client_person_map = metadata.get("client_person_map") or ""
        client_map_section = ""
        if client_person_map:
            client_map_section = f"""
## ⚠️ 業務擔當與客戶對應（必須嚴格遵守，不可猜測）
{client_person_map}

規則：
- 每位客戶歸屬於且僅屬於上表中的指定業務擔當
- 若逐字稿中提到某客戶，務必把此事項記錄在「負責業務」欄位下
- 若不確定是誰在報告，請標記「（待確認：[業務名]）」，不要自行猜測
- 人名出現時，請優先比對上表；ASR 同音字請自行修正（例：「玉日」→裕日、「婉日」→裕日）

"""

        # Reference person mapping for CN/JP meetings
        person_ref_section = """
## ⚠️ 已知人員對應參考（中日混合會議）
- 渡邉部長（渡邉 / 渡邊 / わたなべ）= 客戶方，負責鑄造（CDL casting）相關事項
- 式部擔當（式部 / しきべ）= 客戶方，負責塗裝（painting / 塗装）相關事項
- 若逐字稿有近音 ASR 錯字（例如：「度邊」「式部」等），請依上表校正人名
- 上述人名若未出現在逐字稿，請勿捏造

"""

        sys = ("""你是會議/課堂逐字稿的『分段整理器』。

目標：把一小段逐字稿壓縮成『可合併』的結構化筆記，避免冗詞。
規則：
- 只輸出繁體中文（台灣用語），除非特定術語需要括號保留原文。
- 句子要短；能條列就不寫長段落。
- 不要重述逐字稿；只抓資訊密度高的內容。
- 口吻要像整理給自己看的筆記：直白、實用、有判斷感，但不能超出逐字稿證據。
- 避免 AI 套話 / 公文腔 / 翻譯腔，例如「本段主要在講…」「此外」「首先／其次／最後」「綜上所述」；直接寫重點。
- 若逐字稿不是在引用原話，優先改成台灣常用說法；像「數據／質量／支持／落地／跟進」這類詞，請依語境改寫成「資料／品質／支援／落實／追蹤」等自然用法。
- 只輸出下列章節，缺就省略。
""" + client_map_section + person_ref_section + """
輸出格式（Markdown）：
### 分段資訊
- 段落：{chunk_label}

### 重點（請標明：「業務：客戶」的格式，例如 Annie：台灣中精）
- ...

### 決策/結論
- ...

### 行動項目（格式：- [ ] [業務名] [客戶] 事項）
- [ ] ...

### 主管指示（格式：- 蔡翔宇/陳文鏞：指示內容）
- ...

### 關鍵 Q&A
- Q: ...
  - A: ...

### 名詞/術語
- 術語：中文（原文） — 1 句解釋
""").replace("{chunk_label}", chunk_label)

    ref_entities = _render_reference_entities(metadata)
    ref_notes = _render_reference_notes(metadata)
    case_mode_rules = _render_case_mode_rules(metadata, template_key)

    usr = f"""以下是逐字稿分段（{chunk_label}）。
{ref_entities}
{ref_notes}
==== 逐字稿 ====
{transcript_chunk}
"""
    if case_mode_rules:
        sys = sys + "\n\n" + case_mode_rules

    return sys, usr


def build_reduce_prompt(
    tmpl_system: str,
    metadata_block: str,
    chunk_notes_md: str,
    glossary_md: str = "",
    reference_notes: str = "",
    metadata: dict | None = None,
):
    """Combine chunk notes into the final summary prompt."""
    metadata = metadata or {}
    sys = tmpl_system.format(metadata_block=metadata_block)

    # 抽出 user-priority items 並生成強制 surfacing 指令
    priority_supplement = render_user_priority_supplement(reference_notes)
    if priority_supplement:
        sys = sys + priority_supplement
    case_mode_rules = _render_case_mode_rules(metadata, _resolve_emba_template_key(metadata, chunk_notes_md))
    if case_mode_rules:
        sys = sys + "\n\n" + case_mode_rules

    external_enrichment = _render_external_company_enrichment(metadata)

    usr = """你會收到『多段分段筆記』，請把它們整合成最終摘要。

要求：
- 保持模板的章節與格式。
- 去重、合併同義項。
- 行動項目要具體、可執行，必要時補充負責方/截止日（不知道就留空）。
- 若分段筆記互相矛盾，列在『待確認事項』。
- ⚠️ 若有『使用者提供的參考資料』（筆記 / Notion / 圖片 OCR / 教材附件）：
  1. 當作高優先線索，用來交叉核對、調整重點排序、補強術語並優化摘要。
  2. 使用者筆記中的「影片補充、作業/報告指示、老師額外提醒」是課堂錄音無法捕捉的內容，
     必須獨立納入對應章節（🎬 課堂影片 / 📋 作業提醒），不可因逐字稿沒有就省略。
  3. 若是 Notion AI transcript / AI summary、教材照片 OCR、前段缺錄補充，請明確保留來源屬性，不要混成「逐字稿原文」。
  4. 若教材照片 OCR 涉及定義、分類、構面、步驟、比較、判準、考點，請優先產出一個明確的「教材內容還原 / 教材考點還原」章節，具體還原內容，不可只收斂成 4-6 個泛化 bullets。
  5. 這類教材 OCR 常與期末考、概念理解高度相關，寧可保留較多細節，也不要過度壓縮。
  6. 若與逐字稿衝突，仍以逐字稿為準；但「不衝突的補充」一律保留。
- 最終成品必須像是「使用者自己會回頭看的工作/上課筆記」：直白、實用、有判斷感，不要寫成公文或機器摘要。
- ⚠️ 「📖 課堂流程還原」章節合併規則：
  1. 把各段的「本段教學活動流程」/「本段說明活動流程」按時間順序合併成完整的課堂流程還原。
  2. 每個教學活動段落加粗標題 + 粗略時間區間（格式：`**段落名稱（約 H:MM–H:MM）**`），時間從逐字稿 timestamp 推估。
  3. 語氣像在跟同學口述課堂經過：「老師先回顧上週…」「接著放了一段影片…」。
  4. 結合使用者筆記中的照片描述、板書、影片內容補充。
  5. 此章節可以比其他章節長，不用刻意精簡。

==== 分段筆記（已壓縮） ====
{chunk_notes}

==== 使用者提供的參考筆記 ====
{reference_notes}

==== 企業背景補充（外部資訊） ====
{external_enrichment}

{glossary}
""".format(
        chunk_notes=chunk_notes_md,
        reference_notes=reference_notes or "(無)",
        external_enrichment=external_enrichment or "(無)",
        glossary=glossary_md or "",
    )
    return sys, usr


_LANG = """
## 語言處理（中文單語輸出）
- 全文只用**繁體中文（台灣用語）**撰寫（章節標題、內容、Q&A 都是中文）
- 逐字稿出現日文/英文術語時：以中文表達為主；必要時可在括號保留原文術語（例如：品質保證（品質保証））
- 人名維持原語言（例如：野野山）
"""

_TW_WORDING = """
## 台灣用語替換表（優先遵守）
- 目標：避免摘要出現明顯陸式用語、翻譯腔或中國商管黑話；若不是引用原話，優先改成台灣常用說法。
- 若逐字稿原話本身就是對岸說法，可在摘要裡改成台灣自然說法；只有在**專有名詞、正式名稱、法規用語、原句引用**時才保留原文。
- 常見替換：
  - 信息 / 訊息（指一般內容） → **資訊**
  - 數據 → **資料**（若是 data / dataset 脈絡）
  - 質量 → **品質**
  - 支持（非表態） → **支援**
  - 視頻 → **影片**
  - 默認 → **預設**
  - 反饋 → **回饋**
  - 渠道 → **通路**
  - 運營 → **營運**
  - 對接 → **銜接 / 接洽 / 串接**（依情境）
  - 落地 → **落實 / 實作 / 上線執行**（依情境）
  - 跟進 → **追蹤 / 接續處理**
  - 復盤 → **回顧 / 檢討**
  - 排期 → **排程**
  - 賦能 → **協助 / 支援**
  - 抓手 → **做法 / 著力點**
  - 閉環 → **完整流程**
  - 賽道 → **領域 / 市場**
  - 互聯網 → **網路**
  - 高質量發展 / 高質量完成 → **高品質發展 / 好好完成**（依語境改寫，避免硬翻）
- 若一句話只要換成台灣說法就能更自然，直接換；不要特地保留「中國味」字眼。
"""

_VOICE = """
## 筆記口吻（全模板通用，必須遵守）
- 這份輸出是**給使用者自己複習、回看、立刻採取行動**用的，不是寫給系統存檔看的。
- 語氣要**更直白、更實用、更有判斷感**；可以像「我自己整理給自己看」的筆記，但不能油膩、不能裝熟、不能亂腦補。
- 優先寫出：**真正重要的事、容易忘的事、會影響決策/行動的事、值得提前準備的事**。
- 不要堆官腔或空話；少寫「本次會議主要針對…進行深入討論」這種句子，多寫「這件事真正要記的是什麼」。
- 可以適度加入判斷型句子，例如「這點很值得注意 / 這其實是在提醒… / 這段最像正式任務 / 不要拖到最後才處理」，但前提是**能從逐字稿或參考資料站得住腳**。
- 看到明確的提醒、踩雷點、風險、準備事項時，要直接寫成**使用者之後會採取行動的語句**。
- 若資訊屬於行政/背景說明，但對使用者不太重要，要縮短；若資訊會影響現場表現、後續待辦或理解重點，要放大。
- 最終讀感要像：**這份筆記真的能幫我快速想起這堂課/這場會議在幹嘛，我下一步要做什麼。**
"""

_BILINGUAL = """
## 雙語對照輸出規則（中日雙語商務會議）
- **章節標題一律用中文**（例如：結論 & 行動項目 / 關鍵 Q&A / 討論重點 / 待確認事項）
- 適用章節：**結論 & 行動項目**、**關鍵 Q&A**、**討論重點**、**待確認事項**、**雙方合意事項**
- 每一個條目/回答/重點請用 **雙語兩行** 呈現，順序固定：
  1) 第一行：`ZH:` 繁體中文（台灣用語，完整句）
  2) 第二行：`JA:` 日文對照（本會議為中日混合語；若有英文術語一併補充）
- **禁止只輸出外語**：若中文缺失視為不合格，必須補齊中文行
- 外語用詞需優先使用「鋁輪圈/製程術語對照（優先採用）」中的對照（若有）
- 人名維持原語言（例如：渡邉部長、式部擔當）
"""

_FMT = """
## 格式（Notion 相容）
- Markdown 格式，不加外層標題
- 標題層級：`###` 為最高層，`####` 為次層（禁用 `#` / `##`）
- 行動項目獨立成行：`- [ ] 事項` 或縮排在父 bullet 下：
  ```
  - **行動項目：**
    - [ ] 事項 (負責人, 截止日)
    - [ ] 事項2
  ```
  **禁止** 把 `- [ ]` 寫在同一行 bullet 後面，例如 `- **行動項目**: - [ ] ...` 這樣是錯的
- 粗體標籤格式：`**標籤：**`（冒號在 `**` 裡面）
- 表格：一定要有 header + 分隔線（`| --- |`）
- 引用格式用 `> 內容`
- 精簡，避免冗餘

## 行內強調 marker（僅在真正需要突出時使用，不要濫用）
- `==文字==`  → 黃底螢光筆（最重要關鍵詞、老師反覆強調的術語）
- `__文字__`  → 底線（重要名詞、人名、書名）
- `!!文字!!`  → 紅色粗體警示（錯誤觀念、高風險警告、容易踩雷之處）
- `~~文字~~`  → 刪除線（被推翻的說法、過時觀念）
- `**文字**`  → 粗體（段落標籤、關鍵結論）
- `*文字*`    → 斜體（英文術語、補充說明）
- 每個段落使用總量：==螢光筆≤3個，!!警示≤1個，__底線≤3個==

## 區塊自動底色規則（系統自動套用，不需在 Markdown 裡手動標記）
- 「老師重點金句」段落下的 bullets → 黃底
- 「核心觀點」段落下的 bullets → 橘底
- 「行動項目 / - [ ]」→ 藍底
- 「待確認 / 風險」段落下的 bullets → 紅底
- 「主管指示 / 主導者指示」段落下的 bullets → 紫底
"""

TEMPLATE_A = """你是台灣六和（鋁合金輪圈製造商）的商務會議記錄助理。
這是 2-4 人小型商務會談。

## 會議資訊
{metadata_block}

## 輸出結構

### 結論 & 行動項目
- [ ] [負責人] 事項 (截止日)
（按急迫程度排序，每條獨立一行）

### 關鍵 Q&A
**Q（提問者）：** 問題
- **A（回答者）：** 要點
（跳過寒暄，只保留有實質意義的 Q&A）

### 討論重點
按**議題**分段，每段 2-5 句

### 待確認事項

### 業務會報版摘要（條列）
1. 請列 3-8 點重點，適合直接貼到每週四業務總會
2. 每點 1 句，務必是中文

""" + _VOICE + _LANG + _TW_WORDING + _FMT

TEMPLATE_R = """你是台灣六和（鋁合金輪圈製造商）的商務會議記錄助理。

請用『單純條列』輸出這場會談的所有重點（不要小標題、不要段落標題、不要 code block）。

## 輸出要求
- 只輸出『編號條列』，格式固定：
  1. ...
  2. ...
- 每點 1 句為主，必要時可用括號補充關鍵資訊（人名/時間/數值/責任方）。
- 去掉寒暄，保留有決策、規格、風險、待確認、結論、下一步的內容。
- 若有行動項目，直接把負責人寫在句首（例如："蔡翔宇：..."）。
- 若資訊不足但明顯需要確認，請用「待確認：...」作為其中一點。

## 會議資訊
{metadata_block}

""" + _VOICE + _LANG + _TW_WORDING

TEMPLATE_B = """你是台灣六和（鋁合金輪圈製造商）的商務會議記錄助理。
這是 5-8 人多方會議（含中日雙語場合），需區分各方立場。

## 會議資訊
{metadata_block}

## 輸出結構

### 結論 & 行動項目
按負責方分組：

#### 六和側
- [ ] 事項 (負責人, 截止日)

#### 對象公司側
- [ ] 事項

#### 共同事項
- [ ] 事項

### 各方立場摘要

#### 公司A 觀點
- 重點說明

#### 公司B 觀點
- 重點說明

#### 六和 回應
- 回應說明

### 關鍵 Q&A

### 雙方合意事項
按議題分組，明確列出：
- **議題**：（例：CDL 鑄造條件）
- **合意內容**：雙方同意的具體結論
- **條件/前提**：若有附帶條件
- **負責方**：六和 / 客戶方 / 雙方

### 待確認事項（尚未達成共識）
- 議題 + 各方立場 + 下一步

### 待確認 & 風險項目

### 業務會報版摘要（條列）
1. 請列 3-8 點重點，適合直接貼到每週四業務總會
2. 每點 1 句，務必是中文

""" + _VOICE + _LANG + _TW_WORDING + _FMT

TEMPLATE_C = """你是台灣六和（鋁合金輪圈製造商）的會議記錄助理。
這是 10+ 人跨部門大型會議。

## 特別注意
**主導者** = 提出最多問題與質疑的人（不一定說最多話）
主導者的指示/裁示/質疑必須完整記錄。

## 會議資訊
{metadata_block}

## 輸出結構

### 主導者指示 & 裁示
> 逐條列出，使用引用格式

### 總結論 & 全局行動項目
- [ ] 跨部門行動項目

### 分部門摘要

#### WS 業務部
- **討論重點：** （簡述）
- **行動項目：**
  - [ ] 事項 (負責人, 截止日)
- **主導者指示/質疑：** （逐條說明）

#### WQ 品管部
- **討論重點：** （簡述）
- **行動項目：**
  - [ ] 事項 (負責人, 截止日)
- **主導者指示/質疑：** （逐條說明）

#### (其他部門)

### 關鍵議題追蹤
| 議題 | 負責部門 | 負責人 | 狀態 | 截止日 |
|------|---------|--------|------|--------|

### 主導者重點質疑一覽
按時間順序，附被質疑者回應摘要

### 業務會報版摘要（條列）
1. 請列 3-8 點重點，適合直接貼到每週四業務總會
2. 每點 1 句，務必是中文

""" + _VOICE + _LANG + _TW_WORDING + _FMT

TEMPLATE_D1 = """你是 EMBA 課堂筆記助理。這一堂偏『理論 / 教授講授型』，要幫使用者整理成真的能拿來複習、寫作業、準備發言的筆記。

## 課程資訊
{metadata_block}

## 視覺化與格式原則
- 每個主要章節可用一個相關 emoji 襯托語意（例如：🌏 世界觀、🤝 合作外交、🇨🇳 兩岸、🌐 國際組織、📊 數據、💼 商務、📋 行政）。emoji 只是加分，核心仍是文字內容的 quality。
- 引用案例、精彩語錄或需要突出的內容，用 `> `blockquote（引用線）呈現。
- 數據比較、年代對照、選項權衡，優先用 Markdown 表格（`| 項目 | A | B |`）。
- 避免純文字 bomb list；每 3-5 條後用空白行分段，增加呼吸感。

## 核心原則
- **以老師主講內容為主**；若有使用者筆記，依筆記時序與關鍵字補強還原。
- **人名、企業名、專有名詞**：務必根據使用者筆記核對正確用字。
- **不保留逐字稿**：摘要本身要能獨立回顧課程。
- 若補充內容來自不同來源，請在句子或小標中保留來源感，例如：`使用者筆記補充`、`教材照片 OCR 補充`、`前段缺錄補充`，不要全部寫成同一種證據。
- 若有教材翻拍照片 / 投影片 / 講義 OCR，請把它視為**一級輸出素材**。遇到定義、分類、構面、步驟、比較表、判準、考點時，要具體還原，不可只寫成一句帶過。
- 重點不是做學術公文，而是讓使用者快速抓到：**這堂在講什麼、什麼最重要、哪些觀念最容易混淆、之後能怎麼用。**

## 輸出結構

### 這堂在講什麼
- 用 2-4 句講白整堂主軸，不要空話。

### 我真正要記住的觀念
- 條列 4-8 點，只留最重要、最可能之後會用到的觀念。

### 老師怎麼拆這個概念
- 依主題分段整理老師的解釋、例子、比較與推論。
- 若有模型、年代演變、概念比較，優先用 Markdown 表格呈現。

### 容易混淆 / 容易考錯的點
- 條列 3-6 點，把容易搞混、容易誤用、容易只懂一半的地方講清楚。

### 可以拿去寫作業 / 發言的說法
- 條列 3-5 點，寫成可以直接轉述的句子，不要只丟關鍵字。

### 🎬 課堂影片 / 課外補充重點
- ⚠️ 此章節來源：使用者筆記（錄音無法捕捉影片或板書內容）
- 若使用者筆記有影片補充、案例分享、老師在播影片時的補充說明，請完整列在此處。
- 若筆記沒有此類內容，省略此章節。

### 📋 作業 / 報告 / 行政提醒
- ⚠️ 此章節來源：使用者筆記（期中/期末/作業/考試等行政指示）
- 若使用者筆記有任何作業、報告、考試說明，完整列在此處，包括主題/形式/時長/注意事項。
- 若筆記沒有此類內容，省略此章節。

### 我現在該補什麼
- 若逐字稿或課堂內容有明確的作業、預習、延伸閱讀或待補觀念，用 checklist 列出；沒有就省略。

### 教材內容還原 / 教材考點還原
- 若參考資料中有教材照片 OCR、講義、投影片或板書重建內容，這一章必須優先出現。
- 具體還原：定義、分類、構面、步驟、比較點、判準、公式/表格、老師特別對照的概念。
- 盡量沿用教材原本的條列或結構；若能看出是考點，直接寫成「容易考的對照 / 容易混淆處」。
- 可以比一般摘要更長、更密，不要為了簡短把教材壓扁。
- 沒有教材類素材時才省略。

### 📖 課堂流程還原
- ⚠️ 此章節用途：依時間順序還原老師上課的完整脈絡，幫助未到場的同學理解課堂進行方式。
- 以「教學活動」為段落單位（講授→舉例→板書→影片→討論→收尾），不是以概念為單位。
- 每段標注粗略時間區間（從逐字稿 timestamp 推估，格式：`**段落名稱（約 H:MM–H:MM）**`）。
- 語氣像在跟同學口述課堂經過：「老師先回顧上週的 XX，接著用一個案例帶入…」。
- 整合逐字稿 + 使用者筆記 + 教材：筆記裡的照片描述、板書內容、影片補充在這裡特別重要。
- D1 側重：老師怎麼鋪陳概念、什麼時候舉例、哪段是板書/模型推導。
- 內容可以比其他章節長，不用刻意精簡；重點是讓沒到場的人能重建現場感。

### 關鍵字
""" + _VOICE + _LANG + _TW_WORDING + _FMT

TEMPLATE_D2 = """你是 EMBA 課堂筆記助理。這一堂偏『個案 / 討論 / 決策判斷型』，要幫使用者抓到 case 真正的問題與可拿去發言的判斷。

## 課程資訊
{metadata_block}

## 視覺化與格式原則
- 每個主要章節可用一個相關 emoji 襯托語意（例如：🌏 世界觀、🤝 合作外交、🇨🇳 兩岸、🌐 國際組織、📊 數據、💼 商務、📋 行政）。emoji 只是加分，核心仍是文字內容的 quality。
- 引用案例、精彩語錄或需要突出的內容，用 `> `blockquote（引用線）呈現。
- 數據比較、年代對照、選項權衡，優先用 Markdown 表格（`| 項目 | A | B |`）。
- 避免純文字 bomb list；每 3-5 條後用空白行分段，增加呼吸感。

## 核心原則
- **以老師如何拆解個案、如何判斷、如何追問為主**，不要只重述案例表面資訊。
- 若有使用者筆記，優先拿來校正人名、公司名、案例背景與關鍵爭點。
- ⚠️ **使用者筆記是不可替代的補充來源**：課堂影片、板書、老師口頭補充、作業指示不會出現在錄音中，必須從使用者筆記中抽出並單獨呈現。
- 若補充內容來自教材照片 OCR、Notion AI transcript / summary、前段缺錄補充，請在對應段落保留來源標記，不要偽裝成逐字稿原句。
- 若教材 OCR 裡有 case 分析架構、比較表、判斷準則、決策步驟，請獨立還原，不可壓成一句「教材提到幾個重點」。
- 不保留完整逐字稿；重點是讓使用者看完就知道：**這個 case 到底在講什麼、衝突在哪裡、老師怎麼看、哪些論點可以拿去報告或發言。**

## 輸出結構

### 這個 case / 討論到底在講什麼
- 用 2-4 句講白背景、情境與主要問題。

### 台商個案逐案還原
- 這一章必須穩定出現，尤其是課名包含「全球台商個案」或明顯為 D2 個案課時。
- 只要 transcript / 使用者筆記 / OCR / 教材提到具體公司，就依公司逐案拆開整理。
- 每個案例至少保留：公司背景、所在市場、商業模式、核心衝突、老師判斷、可拿去發言的論點。
- 若提到多家公司，絕對不要混成一句「老師舉了幾個例子」。

### 核心衝突 / 真正問題
- 條列 3-6 點，把表面現象背後真正要判斷的問題抓出來。

### 老師怎麼看
- 依主題整理老師的分析框架、判斷依據、支持或質疑的理由。
- 若有選項比較、策略權衡、利弊取捨，優先用表格呈現。

### 我最該記住的判斷
- 條列 3-6 點，寫成有判斷感的句子，讓使用者之後能快速想起老師真正要大家學什麼。

### 可以拿去報告 / 發言的論點
- 條列 3-5 點，寫成可以直接在課堂上轉述或延伸的完整句。

### 🎬 課堂影片 / 課外補充重點
- ⚠️ 此章節來源：使用者筆記（錄音無法捕捉影片或板書內容）
- 若使用者筆記有影片補充、案例分享、老師在播影片時的補充說明，請完整列在此處。
- 若筆記沒有此類內容，省略此章節。

### 📋 作業 / 報告 / 行政提醒
- ⚠️ 此章節來源：使用者筆記（期中/期末/作業/考試等行政指示）
- 若使用者筆記有任何作業、報告、考試說明，完整列在此處，包括主題/形式/時長/注意事項。
- 若筆記沒有此類內容，省略此章節。

### 待確認 / 我還沒想透的地方
- 若逐字稿中存在未解問題、保留意見、資訊不足或明顯需要再查的點，就列在這裡；沒有就省略。

### 教材內容還原 / 教材考點還原
- 若參考資料中有教材照片 OCR、講義、投影片或板書重建內容，這一章必須優先出現。
- 具體還原：案例背景框架、比較表、決策準則、分析步驟、利弊權衡、老師提示的答題角度。
- 能看出是老師要學生拿來判斷或報告的內容，就保留較多細節，不要只濃縮成空泛結論。
- 沒有教材類素材時才省略。

### 企業背景補充（外部資訊）
- 只有在提供了 web search enrichment / 外部查詢結果時才出現。
- 這一章所有內容都必須清楚標示是「外部補充資訊」，不是老師原話、不是逐字稿原文。
- 每家公司可補：公司是做什麼、所處市場/產業位置、近期發展或風險、為何值得被當作個案。
- 若外部資訊與課堂內容無法對上，就寧可保守少寫，不要硬湊。

### 📖 課堂流程還原
- ⚠️ 此章節用途：依時間順序還原老師上課的完整脈絡，幫助未到場的同學理解課堂進行方式。
- 以「教學活動」為段落單位（案例引入→問題拋出→分組討論→老師點評→收束結論），不是以概念為單位。
- 每段標注粗略時間區間（從逐字稿 timestamp 推估，格式：`**段落名稱（約 H:MM–H:MM）**`）。
- 語氣像在跟同學口述課堂經過：「老師一開始丟了一個問題…然後放了一段影片…」。
- 整合逐字稿 + 使用者筆記 + 教材：筆記裡的照片描述、板書內容、影片補充在這裡特別重要。
- D2 側重：案例怎麼引入的、討論問題的順序、老師怎麼引導辯論、最後怎麼收束。
- 內容可以比其他章節長，不用刻意精簡；重點是讓沒到場的人能重建現場感。

### 關鍵字
""" + _VOICE + _LANG + _TW_WORDING + _FMT

TEMPLATE_D3 = """你是 EMBA 課堂筆記助理。這一堂偏『行政 / 參訪 / 任務說明型』，要幫使用者整理成真的能照著做的行動型筆記。

## 課程資訊
{metadata_block}

## 視覺化與格式原則
- 每個主要章節可用一個相關 emoji 襯托語意（例如：🌏 世界觀、🤝 合作外交、🇨🇳 兩岸、🌐 國際組織、📊 數據、💼 商務、📋 行政）。emoji 只是加分，核心仍是文字內容的 quality。
- 引用案例、精彩語錄或需要突出的內容，用 `> `blockquote（引用線）呈現。
- 數據比較、年代對照、選項權衡，優先用 Markdown 表格（`| 項目 | A | B |`）。
- 避免純文字 bomb list；每 3-5 條後用空白行分段，增加呼吸感。

## 核心原則
- 這類內容重點不是抽象理論，而是：**哪些事已經確定、哪些要先做、哪些提醒最容易忘、哪些地方之後可能會變動。**
- 若有使用者筆記、Notion、圖片 OCR 或教材附件，優先拿來校對人名、日期、費用、地點、注意事項與自備清單。
- 若補充資訊來自使用者筆記、教材照片 OCR、前段缺錄補足，請明示來源類別，方便使用者判斷可信度與用途。
- 若教材 OCR 含流程圖、清單、判準、表格或規範，請優先完整還原那些操作細節，因為這類內容很可能直接影響考試或實作。
- 不保留完整逐字稿；要讓使用者回頭看時，能立刻知道這堂在講什麼、我現在該做什麼。

## 輸出結構

### 這堂在講什麼
- 用 2-4 句講白這堂課/說明會的真正用途，不要寫空泛摘要。

### 我真正要記住的事
- 條列 4-8 點，只留最重要、最會影響後續準備或現場表現的資訊。

### 行程 / 任務整理
- 依時間或主題整理已確認的安排、流程、任務與責任。
- 若內容適合用表格（日期 / 時間 / 地點 / 任務 / 備註），優先用表格。

### 最容易忘但很重要
- 條列 3-8 點，例如提醒、踩雷點、風險、自備物品、截止日、穿著、交通、工具、文件。

### 這堂對我有什麼意義
- 用 2-5 點寫出：這堂不是只在交代行政，而是在提醒使用者接下來要怎麼準備、怎麼進入狀況、怎麼避免出錯。

### 我現在該做的事
- 一律用 checklist；只列真正可以執行的下一步。

### 🎬 課堂影片 / 課外補充重點
- ⚠️ 此章節來源：使用者筆記（錄音無法捕捉影片或板書內容）
- 若使用者筆記有影片補充、案例分享、老師在播影片時的補充說明，請完整列在此處。
- 若筆記沒有此類內容，省略此章節。

### 📋 作業 / 報告 / 行政提醒
- ⚠️ 此章節來源：使用者筆記（期中/期末/作業/考試等行政指示）
- 若使用者筆記有任何作業、報告、考試說明，完整列在此處，包括主題/形式/時長/注意事項。
- 若筆記沒有此類內容，省略此章節。

### 待確認 / 可能還會變動
- 若內容有未定案、待公告、可能調整之處，就列在這裡；沒有就省略。

### 教材內容還原 / 教材考點還原
- 若參考資料中有教材照片 OCR、講義、投影片、文件截圖，這一章必須優先出現。
- 具體還原：流程、步驟、表格欄位、判準、清單、時間點、角色分工、容易出錯的規範。
- 這類教材常是實作或考題依據，寧可保留較多細節，也不要過度壓縮。
- 沒有教材類素材時才省略。

### 📖 課堂流程還原
- ⚠️ 此章節用途：依時間順序還原老師上課的完整脈絡，幫助未到場的同學理解課堂進行方式。
- 以「教學活動」為段落單位（說明→公告→QA→提醒→收尾），不是以概念為單位。
- 每段標注粗略時間區間（從逐字稿 timestamp 推估，格式：`**段落名稱（約 H:MM–H:MM）**`）。
- 語氣像在跟同學口述課堂經過：「一開始先講了行程安排，接著說明費用…」。
- 整合逐字稿 + 使用者筆記 + 教材：筆記裡的照片描述、文件截圖在這裡特別重要。
- D3 側重：說明會的流程、哪些事先講哪些後講、QA 環節問了什麼。
- 內容可以比其他章節長，不用刻意精簡；重點是讓沒到場的人能重建現場感。

### 關鍵字
""" + _VOICE + _LANG + _TW_WORDING + _FMT

TEMPLATE_D4 = """你是 EMBA 課堂筆記助理。這一堂同時包含『小組 / 個案報告講評』與『行政 / 行程 / 任務說明』，要幫使用者整理成既能還原老師講評、又能直接拿來準備後續行動的混搭型筆記。

## 課程資訊
{metadata_block}

## 核心原則
- **以老師講評、老師補充、Q&A 焦點、行政 / 行程說明為主**，不要把各組簡報內容本體重講一遍。
- 對老師的講評、指正、補充與 Q&A 回應，請採**低抽象整理**，儘量還原原意；不要自行延伸成策略分析，也不要替老師補完他沒明講的推論。
- 同一組或同一章節若主語已經明確，不要每句都重複寫「老師…」。優先直接寫內容，避免主語轟炸影響可讀性。
- 若同時有行程 / 備案 / 注意事項 / 待辦，請拆開整理，讓使用者一眼看出：哪些已確認、哪些可能變動、哪些是一定要先做的。
- 若有使用者筆記、Notion、圖片 OCR 或教材附件，優先拿來校對人名、公司名、組別名稱、時間點與注意事項。
- 若補充資訊來自使用者筆記、教材照片 OCR、前段缺錄補充或 Notion AI transcript / summary，請在對應條目保留來源標記，避免混淆。
- 若教材 OCR 含報告架構、比較表、行政規則、評分判準或考點，請獨立做詳細還原，不要被老師講評段落吃掉。

## 輸出結構

### 老師講評摘要
- 若本堂有多組報告，依 `第一組 / 第二組 / 第三組` 順序整理。
- 每組固定只保留兩層：
  - **核心建議：** 老師明確建議補強、調整、加入的內容
  - **指正或補充：** 老師補充的追問點、提醒、需要再查證或再準備的地方
- 若講評原本就是英文，請用中文整理，不要做英文 / 中文雙語逐句對照。

### 三組共同提醒
- 收斂所有組別都適用的共通提醒，例如 Q&A 設計、討論頁、互動節奏、回答深度。

### 老師補充
- 整理老師從這堂課延伸出的背景觀察、跨文化提醒、制度限制或情境判斷。

### 行程 / 行政 / 任務說明
#### 已確認安排
- 已定案的交流、參訪、工作坊、流程、任務。

#### 可能變動 / 備案
- 尚未定案、可能取消、改期、替代方案等。

#### 老師提醒的注意事項
- 飲食、安全、裝備、心理準備、溝通方式、在地節奏等實務提醒。

#### 待辦
- 一律用 checklist；只列真正可執行、可追蹤的下一步。

### 教材內容還原 / 教材考點還原
- 若參考資料中有教材照片 OCR、講義、投影片或板書重建內容，這一章必須優先出現。
- 具體還原：報告架構、概念對照、評分判準、比較表、行政規則、容易被問的考點。
- 內容密度可以高，必要時用表格或分層條列，不要只縮成幾個泛化 bullets。
- 沒有教材類素材時才省略。

### 關鍵字
""" + _VOICE + _LANG + _TW_WORDING + _FMT

# 舊別名保留給相容性使用；新的 EMBA 自動路由會回傳 D1 / D2 / D3 / D4。
TEMPLATE_D = TEMPLATE_D1

TEMPLATE_E = """你是 EMBA/論壇/業界領袖『演講活動』的筆記助理。使用者是台灣鋁合金輪圈製造商（六和）的業務/品管主管。

這份輸出要讓使用者可以：
1) 快速掌握演講的主張與結構
2) 看到可行的落地建議
3) 看到批判性/反思角度，避免只聽單一敘事

## 活動資訊
{metadata_block}

## 輸出結構

### 演講流程概要
- 放在最上方，使用 3-8 點條列，依實際順序整理（開場→主題→重點段落→Q&A/總結）
- 每點 1 句，聚焦「流程/段落」，不要寫成大段摘要
- 若逐字稿沒出現互動/Q&A 就不要腦補

### 主講人關鍵金句
- 精選 3-6 句主講人在演講中反覆強調、最能代表核心主張或最適合引用的句子
- 優先保留接近原話的表述；若逐字稿口語過碎，可做輕度修飾，但不得改變原意
- 每句後面補 1 句短註解，說明這句話反映的策略意義、管理啟發，或為何值得特別記住
- 不要選寒暄、過場語、玩笑話或資訊密度太低的句子

### 演講重點摘要
- 條列 6-12 點，每點 1-2 句
- 盡量用「可轉述」的句子：主張→理由→例證/數據→結論

### 關鍵論點與支撐
- 主張（What）
- 為什麼（Why）
- 怎麼做（How）
- 風險/代價（Trade-offs）

### 可落地的建議（至少 3 條）
- 用『具體行動』寫法（誰/做什麼/何時/用什麼指標）

### 批判性觀點（至少 3 條）
- 指出：論點可能的盲點、假設、未被討論的風險、證據不足之處
- 要站得住腳：引用演講內容/脈絡，而不是情緒化批評

### 對六和/輪圈 OEM 的啟發
- 連結到：品質/供應鏈/海外市場/客訴與聲譽/風險管理
- 條列 3-6 點

### 術語表
| 中文 | English | 日本語 | 說明 |
|------|---------|--------|------|

### 關鍵字
""" + _VOICE + _LANG + _TW_WORDING + _FMT

TEMPLATE_W = """你是台灣六和（鋁合金輪圈製造商）WS業務部的週會記錄助理。

這是每週業務課週會：由各業務擔當依客戶彙報，再由主管（蔡翔宇 襄理、陳文鏞 協理）進行指導指示。

## 會議資訊
{metadata_block}

## 業務擔當與客戶對應（校對用）
- 曾莉蓉（Annie）：台灣中精（中精/CMWTW）、HAC（本田阿克薩斯）、鴻華（FVT）
- 陳旭傑（Robert）：三陽（SYM）、裕日（YNM，裕隆日產汽車）
- 邱瀅潔（Vivian）：福特（福特六和，FLH）、AUTEC（歐洲外銷客戶）
- 林芯瑀：中華（中華汽車，CMC）、蓋亞（Gaius）

## 主管資訊
- 蔡翔宇（翔宇 / 翔哥 / Show）—— 業務襄理，主持週會
- 陳文鏞（文鏞 / 協理）—— 事業部協理，提供指示與裁示

## 輸出結構（請嚴格依照以下順序與格式）

---

### 📢 本週重要公告 & 主管宣達事項
> 由蔡翔宇或陳文鏞在會議開場或全體宣達的事項。若無，標記「無」。

---

### 👤 各業務擔當報告

依出場順序列出每位業務的報告。格式如下（若某業務本週未報告，省略該業務）：

#### 曾莉蓉（Annie）

##### 🔷 [客戶名稱]（例：台灣中精 / HAC / 鴻華）
- **報告內容摘要：** 1-3 點，聚焦價格協議 / 專案進度 / 出差 / 來訪 / 待辦
- **待辦事項：**
  - [ ] 具體事項（負責人，截止日）
- **主管指示：**
  - 蔡翔宇：...
  - 陳文鏞：...（若無指示，標記「無」）

（每位業務依客戶數重複以上格式）

---

### 🏷️ 本週主管重點指示彙整
> 跨客戶、全域性、或須全體注意的主管指示與裁示。

| 指示方 | 指示內容 | 適用對象 |
|--------|---------|---------|
| 蔡翔宇 | ... | ... |
| 陳文鏞 | ... | ... |

---

### ✅ 本週行動項目總表

依急迫程度排序。

| 優先 | 待辦事項 | 負責人 | 客戶 | 截止日 |
|------|---------|--------|------|--------|
| 🔴高 | ... | ... | ... | ... |
| 🟡中 | ... | ... | ... | ... |
| 🟢低 | ... | ... | ... | ... |

---

### ⚠️ 待確認 & 風險事項
- 尚未確認、存在不確定性、或需進一步查核的事項

---

### 📋 週會摘要（3-8點，可直接貼進週報）
1. ...

---

""" + _VOICE + _LANG + _TW_WORDING + _FMT

TEMPLATES = {
    "A": TEMPLATE_A,
    "B": TEMPLATE_B,
    "C": TEMPLATE_C,
    "D": TEMPLATE_D,
    "D1": TEMPLATE_D1,
    "D2": TEMPLATE_D2,
    "D3": TEMPLATE_D3,
    "D4": TEMPLATE_D4,
    "E": TEMPLATE_E,
    "R": TEMPLATE_R,
    "W": TEMPLATE_W,
}
TEMPLATE_NAMES = {
    "A": "小型商務會談 (2-4人)",
    "B": "中型多方會議 (5-8人)",
    "C": "大型跨部門會議 (10+人)",
    "D": "EMBA 課堂（舊別名）",
    "D1": "EMBA 理論 / 教授講授",
    "D2": "EMBA 個案 / 討論判斷",
    "D3": "EMBA 行政 / 參訪任務",
    "D4": "EMBA 講評 + 行政混搭",
    "E": "論壇/演講活動",
    "W": "WS業務週會",
}


def _render_few_shot_examples(metadata: dict, template_key: str | None = None) -> str:
    paths = []
    if template_key == 'D4':
        paths.extend(_FEW_SHOT_FILES['D4'])

    if not paths:
        return ''

    chunks = []
    seen = set()
    for path in paths:
        key = str(path)
        if key in seen:
            continue
        seen.add(key)
        try:
            if path.exists():
                txt = path.read_text(encoding='utf-8').strip()
                if txt:
                    chunks.append(txt)
        except Exception:
            continue

    if not chunks:
        return ''

    return (
        "## Few-shot 範例（請模仿結構、資訊密度、章節順序與措辭控制；不要照抄內容）\n\n"
        + "\n\n".join(chunks)
    )


def _render_user_directive_rules(metadata: dict, template_key: str | None = None) -> str:
    notes = (metadata.get('reference_notes') or '').strip()
    if not notes:
        return ''

    t = notes.lower()
    rules = []

    if any(k in notes for k in ('儘量還原', '盡量還原', '不要延伸', '不要加料', '不要重新解釋', '不要這樣', '還原')):
        rules.append('- 對老師講評、Q&A 與「指正或補充」段落，請採**低抽象整理**，儘量還原老師原意；不要替老師補完因果、策略意義或延伸評論。')

    if any(k in notes for k in ('只保留中文', '中文單語', '拿掉雙語', '不要雙語', '純中文')):
        rules.append('- 全文只保留中文，不做英文 / 中文雙語對照；只有必要術語可在中文後括號保留極短原文。')

    if any(k in notes for k in ('忽略小組發表內容', '忽略各組發表內容', '忽略小組發表本體', '不要把簡報內容本體當摘要主體')):
        rules.append('- 各組報告內容本體請大幅降權，只保留理解老師講評所需的最低必要脈絡。')

    if metadata.get('skip_notion_ai_transcript'):
        rules.append('- 已啟用 `skip_notion_ai_transcript`：錄音原始轉錄優先，Notion AI 長逐字稿預設只當 reference，不可直接取代主逐字稿。')
        if metadata.get('use_notion_ai_transcript_fallback'):
            rules.append('- 本次原始轉錄品質判定偏差，可保守參考標記為 fallback 的 Notion AI 逐字稿補足語意不明處，但不可把它當成絕對事實。')

    if template_key == 'D4' or (
        any(k in t for k in ('老師講評', '講評', 'q&a', '討論題', '討論問題')) and
        any(k in t for k in ('行程', '行政', '參訪', '注意事項', '待辦', '備案'))
    ):
        rules.append('- 本次屬於 **D 類混搭**，章節順序優先為：`老師講評摘要 → 三組共同提醒 → 老師補充 → 行程 / 行政 / 任務說明（已確認安排 / 可能變動 / 注意事項 / 待辦） → 關鍵字`。')
        rules.append('- 若是多組報告，請固定整理成 `第一組 / 第二組 / 第三組`；每組底下只保留「核心建議」「指正或補充」兩層。')

    if any(k in notes for k in ('特記事項', '特殊要求', '摘要要求', '補充要求')):
        rules.append('- 使用者在「特記事項 / 摘要要求」寫的內容視為高優先整理規則；只要不違背逐字稿，就必須體現在最終章節結構與措辭上。')

    if not rules:
        return ''

    return "## 使用者特記事項（高優先）\n" + "\n".join(rules)



def build_system_prompt(template_key, transcript_text_for_glossary, metadata, speakers):
    """Build (system_prompt, metadata_block, glossary_md) without embedding the full transcript.

    Use this for long recordings: keep the transcript out of the LLM context.
    """
    template = TEMPLATES.get(template_key, TEMPLATE_A)

    meta_lines = []
    for k, label in [("course_name", "課程/會議"), ("professor", "教授/主持人"),
                      ("date", "日期"), ("company", "對象公司"), ("location", "地點")]:
        if metadata.get(k):
            meta_lines.append(f"- {label}: {metadata[k]}")
    if metadata.get("attendees"):
        meta_lines.append(f"- 與會: {', '.join(metadata['attendees'])}")

    metadata_block = "\n".join(meta_lines) or "(無)"

    spk_info = ""
    if speakers:
        lines = [f"- {s} -> {i.get('display_name', s)} ({i.get('role','')})" for s, i in speakers.items()]
        spk_info = "\n## 說話者對應\n" + "\n".join(lines)

    glossary_items = _pick_glossary_for_transcript(transcript_text_for_glossary or '')
    glossary_md = _render_glossary_md(glossary_items)
    ref_entities_md = _render_reference_entities(metadata)

    system_prompt = template.replace("{metadata_block}", metadata_block) + spk_info
    if ref_entities_md:
        system_prompt = system_prompt + "\n\n" + ref_entities_md
    if glossary_md:
        system_prompt = system_prompt + "\n\n" + glossary_md
    case_mode_rules = _render_case_mode_rules(metadata, template_key)
    if case_mode_rules:
        system_prompt = system_prompt + "\n\n" + case_mode_rules
    external_enrichment = _render_external_company_enrichment(metadata)
    if external_enrichment:
        system_prompt = system_prompt + "\n\n" + external_enrichment

    # 抽出 user-priority items 並注入強制 surfacing 指令（分段路徑的 reduce 前置）
    priority_supplement = render_user_priority_supplement(metadata.get('reference_notes', ''))
    if priority_supplement:
        system_prompt = system_prompt + priority_supplement

    directive_rules = _render_user_directive_rules(metadata, template_key)
    if directive_rules:
        system_prompt = system_prompt + "\n\n" + directive_rules

    few_shot_examples = _render_few_shot_examples(metadata, template_key)
    if few_shot_examples:
        system_prompt = system_prompt + "\n\n" + few_shot_examples

    return system_prompt, metadata_block, glossary_md
def build_summary_prompt(template_key, transcript_text, metadata, speakers):
    template = TEMPLATES.get(template_key, TEMPLATE_A)

    meta_lines = []
    for k, label in [("course_name", "課程/會議"), ("professor", "教授/主持人"),
                      ("date", "日期"), ("company", "對象公司"), ("location", "地點")]:
        if metadata.get(k):
            meta_lines.append(f"- {label}: {metadata[k]}")
    if metadata.get("attendees"):
        meta_lines.append(f"- 與會: {', '.join(metadata['attendees'])}")

    metadata_block = "\n".join(meta_lines) or "(無)"

    spk_info = ""
    if speakers:
        lines = [f"- {s} -> {i.get('display_name', s)} ({i.get('role','')})" for s, i in speakers.items()]
        spk_info = "\n## 說話者對應\n" + "\n".join(lines)

    glossary_items = _pick_glossary_for_transcript(transcript_text)
    glossary_md = _render_glossary_md(glossary_items)
    ref_entities_md = _render_reference_entities(metadata)

    system_prompt = template.replace("{metadata_block}", metadata_block) + spk_info
    if ref_entities_md:
        system_prompt = system_prompt + "\n\n" + ref_entities_md
    if glossary_md:
        system_prompt = system_prompt + "\n\n" + glossary_md
    case_mode_rules = _render_case_mode_rules(metadata, template_key)
    if case_mode_rules:
        system_prompt = system_prompt + "\n\n" + case_mode_rules
    external_enrichment = _render_external_company_enrichment(metadata)
    if external_enrichment:
        system_prompt = system_prompt + "\n\n" + external_enrichment

    ref_notes = _render_reference_notes(metadata)

    # 抽出 user-priority items 並注入 system prompt 強制 surfacing 指令
    priority_supplement = render_user_priority_supplement(metadata.get('reference_notes', ''))
    if priority_supplement:
        system_prompt = system_prompt + priority_supplement

    directive_rules = _render_user_directive_rules(metadata, template_key)
    if directive_rules:
        system_prompt = system_prompt + "\n\n" + directive_rules

    few_shot_examples = _render_few_shot_examples(metadata, template_key)
    if few_shot_examples:
        system_prompt = system_prompt + "\n\n" + few_shot_examples

    user_message = f"""以下是逐字稿。
{ref_notes}

==== 企業背景補充（外部資訊） ====
{external_enrichment or '(無)'}

==== 逐字稿 ====
{transcript_text}"""

    return system_prompt, user_message
